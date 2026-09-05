"""Structured records -> a formatted .xlsx workbook.

Every extraction stage produces the same shape underneath: a list of dicts, one per line
item or transaction, one dict key per column. This is the one place that turns those into
a workbook a fund manager can open -- column widths, headers and sheet-per-document-type
live here once, rather than wherever a stage happens to finish.
"""
from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

HEADER_FONT = Font(bold=True)
SHEET_NAME_LIMIT = 31  # Excel's own ceiling; longer names are silently rejected otherwise.


def write_workbook(sheets: dict[str, list[dict[str, Any]]], destination: Path) -> Path:
    """One sheet per key in `sheets`, e.g. {"Balance Sheet": [...], "Income Statement": [...]}.

    Column order follows the first record's key order, because dict insertion order is
    the only ordering a plain list of dicts carries. Every record in a sheet is expected
    to share the same keys -- a stage that produces ragged records should pad them before
    calling this, rather than leave it to guess at a header row.
    """
    book = Workbook()
    book.remove(book.active)
    for name, records in sheets.items():
        sheet = book.create_sheet(title=name[:SHEET_NAME_LIMIT])
        if not records:
            continue
        headers = list(records[0].keys())
        sheet.append(headers)
        for cell in sheet[1]:
            cell.font = HEADER_FONT
        for record in records:
            sheet.append([record.get(header) for header in headers])
        for index, header in enumerate(headers, start=1):
            widest = max([len(str(header))] + [len(str(r.get(header, ""))) for r in records])
            sheet.column_dimensions[get_column_letter(index)].width = min(widest + 2, 60)
    destination.parent.mkdir(parents=True, exist_ok=True)
    book.save(destination)
    return destination
