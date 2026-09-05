"""The extraction primitives, checked against a real statement PDF and a round trip
through a real workbook -- not mocks, because a mock PDF cannot tell us pdfplumber's
table layout assumptions still hold.

Run:  python -m pytest tests/ -q      (or: python tests/test_extraction.py)
"""
from __future__ import annotations

import sys
import tempfile
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from openpyxl import load_workbook

from src.extraction import pdf_text, workbook_writer
from src.spine.build import STATEMENTS


def test_extract_reads_every_page_of_a_real_statement() -> None:
    one_statement = next(iter(sorted(STATEMENTS.glob("*.pdf"))))
    document = pdf_text.extract(one_statement)

    assert document.pages, "expected at least one page"
    assert "Statement details" in document.text
    # Every page of a bank statement carries its transaction table.
    assert all(page.tables for page in document.pages)


def test_write_workbook_round_trips() -> None:
    sheets = {
        "Balance Sheet": [
            {"Line item": "Cash", "Amount": 1000},
            {"Line item": "Investments", "Amount": 5000},
        ],
        "Income Statement": [{"Line item": "Management fee", "Amount": -50}],
    }
    with tempfile.TemporaryDirectory() as tmp:
        destination = Path(tmp) / "extracted.xlsx"
        workbook_writer.write_workbook(sheets, destination)

        book = load_workbook(destination)
        assert book.sheetnames == ["Balance Sheet", "Income Statement"]
        rows = list(book["Balance Sheet"].iter_rows(values_only=True))
        assert rows[0] == ("Line item", "Amount")
        assert rows[1] == ("Cash", 1000)


def test_write_workbook_skips_empty_sheets_without_crashing() -> None:
    with tempfile.TemporaryDirectory() as tmp:
        destination = Path(tmp) / "extracted.xlsx"
        workbook_writer.write_workbook({"Cash Flow": []}, destination)
        assert destination.exists()


if __name__ == "__main__":
    test_extract_reads_every_page_of_a_real_statement()
    test_write_workbook_round_trips()
    test_write_workbook_skips_empty_sheets_without_crashing()
    print("all extraction checks pass")
